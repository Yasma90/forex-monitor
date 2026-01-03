from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional
import csv
import io
import json
from datetime import datetime

from ...models.database import get_db
from ...models.exchange import ExchangeRateResponse, ExchangeRateCreate, ExchangeRateHistoryResponse
from ...services.exchange import ExchangeRateFetcher, ExchangeRateRepository

router = APIRouter(prefix="/api/exchange", tags=["exchange"])


@router.get("/rate", response_model=ExchangeRateResponse)
async def get_current_rate(
    base: str = Query(default="USD", description="Base currency code"),
    target: str = Query(default="EUR", description="Target currency code"),
    db: AsyncSession = Depends(get_db)
):
    """
    Get the current exchange rate for a currency pair.
    Includes 24h change if historical data is available.
    """
    repo = ExchangeRateRepository(db)

    # Get latest rate from DB
    latest = await repo.get_latest(base, target)

    # If no recent rate (older than 30 min), fetch fresh
    if not latest or _is_stale(latest.timestamp):
        fetcher = ExchangeRateFetcher()
        try:
            fresh_data = await fetcher.fetch_rate(base, target)
            if fresh_data:
                rate_create = ExchangeRateCreate(
                    base_currency=fresh_data["base_currency"],
                    target_currency=fresh_data["target_currency"],
                    rate=fresh_data["rate"],
                    source=fresh_data["source"]
                )
                latest = await repo.save_rate(rate_create)
        finally:
            await fetcher.close()

    if not latest:
        raise HTTPException(status_code=503, detail="Unable to fetch exchange rate")

    # Calculate 24h change
    rate_24h_ago = await repo.get_rate_24h_ago(base, target)
    change_24h = None
    change_percent_24h = None

    if rate_24h_ago:
        change_24h = latest.rate - rate_24h_ago.rate
        change_percent_24h = (change_24h / rate_24h_ago.rate) * 100

    return ExchangeRateResponse(
        id=latest.id,
        base_currency=latest.base_currency,
        target_currency=latest.target_currency,
        rate=latest.rate,
        source=latest.source,
        timestamp=latest.timestamp,
        change_24h=change_24h,
        change_percent_24h=change_percent_24h
    )


@router.get("/history", response_model=ExchangeRateHistoryResponse)
async def get_rate_history(
    base: str = Query(default="USD", description="Base currency code"),
    target: str = Query(default="EUR", description="Target currency code"),
    days: int = Query(default=30, ge=1, le=365, description="Number of days of history"),
    db: AsyncSession = Depends(get_db)
):
    """
    Get historical exchange rates for a currency pair.
    """
    repo = ExchangeRateRepository(db)

    # Get historical rates from DB
    rates = await repo.get_history(base, target, days)

    # If no data, try to fetch historical data
    if not rates:
        fetcher = ExchangeRateFetcher()
        try:
            historical = await fetcher.fetch_historical(base, target, days)
            if historical:
                # Save to DB for future use
                for rate_data in historical:
                    rate_create = ExchangeRateCreate(
                        base_currency=rate_data["base_currency"],
                        target_currency=rate_data["target_currency"],
                        rate=rate_data["rate"],
                        source="frankfurter"
                    )
                    await repo.save_rate(rate_create)
                rates = await repo.get_history(base, target, days)
        finally:
            await fetcher.close()

    if not rates:
        raise HTTPException(status_code=404, detail="No historical data available")

    # Calculate stats
    stats = await repo.get_stats(base, target, days)

    return ExchangeRateHistoryResponse(
        rates=[ExchangeRateResponse(
            id=r.id,
            base_currency=r.base_currency,
            target_currency=r.target_currency,
            rate=r.rate,
            source=r.source,
            timestamp=r.timestamp
        ) for r in rates],
        min_rate=stats["min_rate"],
        max_rate=stats["max_rate"],
        avg_rate=stats["avg_rate"],
        period_days=stats["period_days"]
    )


@router.post("/refresh")
async def refresh_rate(
    base: str = Query(default="USD"),
    target: str = Query(default="EUR"),
    db: AsyncSession = Depends(get_db)
):
    """
    Force refresh the exchange rate (bypasses cache).
    """
    fetcher = ExchangeRateFetcher()
    repo = ExchangeRateRepository(db)

    try:
        fresh_data = await fetcher.fetch_rate(base, target)
        if fresh_data:
            rate_create = ExchangeRateCreate(
                base_currency=fresh_data["base_currency"],
                target_currency=fresh_data["target_currency"],
                rate=fresh_data["rate"],
                source=fresh_data["source"]
            )
            rate = await repo.save_rate(rate_create)
            return {"message": "Rate refreshed", "rate": rate.rate, "source": rate.source}
    finally:
        await fetcher.close()

    raise HTTPException(status_code=503, detail="Failed to refresh rate")


@router.get("/export/csv")
async def export_to_csv(
    base: str = Query(default="USD", description="Base currency code"),
    target: str = Query(default="EUR", description="Target currency code"),
    days: int = Query(default=30, ge=1, le=365, description="Number of days of history"),
    db: AsyncSession = Depends(get_db)
):
    """
    Export historical exchange rates to CSV file.
    """
    repo = ExchangeRateRepository(db)
    rates = await repo.get_history(base, target, days)

    if not rates:
        raise HTTPException(status_code=404, detail="No historical data available")

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(["Date", "Base", "Target", "Rate", "Source"])

    # Data rows
    for rate in rates:
        writer.writerow([
            rate.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            rate.base_currency,
            rate.target_currency,
            rate.rate,
            rate.source
        ])

    output.seek(0)

    # Generate filename with date
    filename = f"forex_{base}_{target}_{days}d_{datetime.now().strftime('%Y%m%d')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/json")
async def export_to_json(
    base: str = Query(default="USD", description="Base currency code"),
    target: str = Query(default="EUR", description="Target currency code"),
    days: int = Query(default=30, ge=1, le=365, description="Number of days of history"),
    db: AsyncSession = Depends(get_db)
):
    """
    Export historical exchange rates to JSON file.
    """
    repo = ExchangeRateRepository(db)
    rates = await repo.get_history(base, target, days)

    if not rates:
        raise HTTPException(status_code=404, detail="No historical data available")

    # Build JSON data
    data = {
        "metadata": {
            "base_currency": base,
            "target_currency": target,
            "period_days": days,
            "exported_at": datetime.now().isoformat(),
            "total_records": len(rates)
        },
        "rates": [
            {
                "date": rate.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                "rate": rate.rate,
                "source": rate.source
            }
            for rate in rates
        ]
    }

    # Generate filename with date
    filename = f"forex_{base}_{target}_{days}d_{datetime.now().strftime('%Y%m%d')}.json"

    return StreamingResponse(
        iter([json.dumps(data, indent=2)]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel")
async def export_to_excel(
    base: str = Query(default="USD", description="Base currency code"),
    target: str = Query(default="EUR", description="Target currency code"),
    days: int = Query(default=30, ge=1, le=365, description="Number of days of history"),
    db: AsyncSession = Depends(get_db)
):
    """
    Export historical exchange rates to Excel file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Excel export not available. Install openpyxl: pip install openpyxl"
        )

    repo = ExchangeRateRepository(db)
    rates = await repo.get_history(base, target, days)

    if not rates:
        raise HTTPException(status_code=404, detail="No historical data available")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{base}_{target}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ["Fecha", "Base", "Target", "Tasa", "Fuente"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row, rate in enumerate(rates, 2):
        ws.cell(row=row, column=1, value=rate.timestamp.strftime("%Y-%m-%d %H:%M:%S")).border = thin_border
        ws.cell(row=row, column=2, value=rate.base_currency).border = thin_border
        ws.cell(row=row, column=3, value=rate.target_currency).border = thin_border
        ws.cell(row=row, column=4, value=rate.rate).border = thin_border
        ws.cell(row=row, column=5, value=rate.source).border = thin_border

    # Auto-adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with date
    filename = f"forex_{base}_{target}_{days}d_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _is_stale(timestamp, max_age_minutes: int = 30) -> bool:
    """Check if a timestamp is older than max_age_minutes"""
    from datetime import datetime, timedelta
    return datetime.utcnow() - timestamp > timedelta(minutes=max_age_minutes)
